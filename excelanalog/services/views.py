import os
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from .forms import CheckListForm, ReestrForm
from .models import Columns, CheckList, Reestr
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from openpyxl.styles import Border, Side


# Create your views here.
def index(request):
    return render(request, 'base.html')


def svod(request):
    if request.method == 'POST':
        files = request.FILES.getlist('files')  # Получение списка загруженных файлов

        # Создание пустого сводного DataFrame для первого листа
        summary_df1 = pd.DataFrame(
            columns=['номер п/п', 'Код КП(общий)', 'Код КП(промежуточный)', 'Наименование ИП', 'Описание КП',
                     'Переодичность проведения', 'Способ подсчета результаты проведения КП',
                     'Подразделение, ответственное за проведение контрольной процедуры', 'Исполнитель КП',
                     'Количество выполненых КП', 'Количество выявленных ошибок'])

        # Создание пустого сводного DataFrame для второго листа
        summary_df2 = pd.DataFrame(columns=['№ п/п', 'Код КП(промежуточный)', 'Исполнитель ИП', 'номер чек листа',
                                            'Объект контроля (договор, акт, счет-фактура, КС-2 и др.)',
                                            'Дата документа', 'Номер документа', 'Количество документов/операций',
                                            'Количество ошибок/нарушений', 'Примечание'])

        summary_df3 = pd.DataFrame(
            columns=['номер п/п', 'Код КП(общий)', 'Код КП(промежуточный)', 'Наименование контрольной процедуры', 'Описание КП',
                     'Переодичность проведения', 'Способ подсчета результаты проведения КП',
                     'Подразделение, ответственное за проведение контрольной процедуры', 'Исполнитель КП',
                     'Количество выполненых КП', 'Количество выявленных ошибок'])



        # Обработка каждого загруженного файла
        for file in files:
            # Чтение первого листа файла и взятие только значений
            df1 = pd.read_excel(file, sheet_name='Sheet', usecols="A:K", header=None, skiprows=8, nrows=1)
            df1 = df1.set_axis(['номер п/п', 'Код КП(общий)', 'Код КП(промежуточный)', 'Наименование ИП', 'Описание КП',
                                'Переодичность проведения', 'Способ подсчета результаты проведения КП',
                                'Подразделение, ответственное за проведение контрольной процедуры', 'Исполнитель КП',
                                'Количество выполненых КП', 'Количество выявленных ошибок', ], axis=1)

            # Получение имени файла без расширения
            file_name = os.path.splitext(file.name)[0]

            # Добавление столбика "Документ" в DataFrame и заполнение его названием файла
            df1['Документ'] = file_name

            df1.drop(['Способ подсчета результаты проведения КП', 'Описание КП', 'Переодичность проведения'], axis=1,
                     inplace=True)
            df1 = df1.reset_index(drop=True)
            df1 = df1.rename_axis([None], axis=1)
            summary_df1 = pd.concat([summary_df1, df1], ignore_index=True)

            filial_value = 'Значение филиала'  # Значение филиала, которое вы хотите добавить

            summary_df1['филиал'] = filial_value



            # Чтение второго листа файла и взятие только значений
            df2 = pd.read_excel(file, sheet_name='Sheet2', usecols="A:J", header=None, skiprows=3, nrows=5)
            df2 = df2.set_axis(['№ п/п', 'Код КП(промежуточный)', 'Исполнитель ИП', 'номер чек листа',
                                            'Объект контроля (договор, акт, счет-фактура, КС-2 и др.)',
                                            'Дата документа', 'Номер документа', 'Количество документов/операций',
                                            'Количество ошибок/нарушений', 'Примечание'], axis=1)

            ########################################
            sum_docs = df2['Количество документов/операций'].sum()
            sum_errors = df2['Количество ошибок/нарушений'].sum()

            # Создаем новый DataFrame с суммами

            ##########################################
            df2 = df2.reset_index(drop=True)
            df2 = df2.rename_axis([None], axis=1)
            summary_df2 = pd.concat([summary_df2, df2], ignore_index=True)



            df3 = pd.read_excel(file, sheet_name='Sheet', usecols="A:K", header=None, skiprows=8, nrows=1)
            df3 = df3.set_axis(
                ['номер п/п', 'Код КП(общий)', 'Код КП(промежуточный)', 'Наименование контрольной процедуры', 'Описание КП',
                     'Переодичность проведения', 'Способ подсчета результаты проведения КП',
                     'Подразделение, ответственное за проведение контрольной процедуры', 'Исполнитель КП',
                     'Количество выполненых КП', 'Количество выявленных ошибок'], axis=1)

            file_name3 = 'Чек-лист/Реестр объектов контроля'

            # Добавление столбика "Документ" в DataFrame и заполнение его названием файла
            df3['Документ, подтверждающий проведение контрольной процедуры'] = file_name3

            df3.drop(['Код КП(промежуточный)', 'Описание КП', 'Переодичность проведения', 'Способ подсчета результаты проведения КП', 'Подразделение, ответственное за проведение контрольной процедуры', 'Исполнитель КП'], axis=1,
                     inplace=True)
            df3 = df3.reset_index(drop=True)
            df3 = df3.rename_axis([None], axis=1)
            summary_df3 = pd.concat([summary_df3, df3], ignore_index=True)


        # Создание нового файла Excel с двумя листами
        with pd.ExcelWriter('summary.xlsx', engine='openpyxl') as writer:
            summary_df1.to_excel(writer, sheet_name='Sheet', index=False)
            summary_df2.to_excel(writer, sheet_name='Sheet2', index=False)
            summary_df3.to_excel(writer, sheet_name='Sheet3', index=False)

            # Получение объекта workbook
            workbook = writer.book

            # Получение объекта worksheet для первого листа
            worksheet1 = writer.sheets['Sheet']
            #worksheet1.merge_cells('J14:J30')

            num_rows = summary_df1.shape[0]

            # Установка границ объединения ячеек





            values = ['АУП', 'Югорское УМТС и К', 'УОВОФ', 'Надымское УАВР', 'Югорское УАВР', 'Белоярское УАВР',
                      'Надымское УТТиСТ', 'Югорское УТТиСТ', 'Белоярское УТТиСТ', 'ИТЦ',
                      'Учебно-производственный центр', 'УЭЗ и С', 'Управление связи', 'Бобровское ЛПУ', 'Верхнеказымское ЛПУ', 'Ивдельское ЛПУ', 'Казымское ЛПУ',
                      'Карпинское ЛПУ', 'Комсомольское ЛПУ', 'Краснотурьинское ЛПУ',
                      'Лонг-Юганское ЛПУ', 'Надымское ЛПУ', 'Нижнетуринское ЛПУ', 'Ново-Уренгойское ЛПУ', 'Ныдинское ЛПУ', 'Октябрьское ЛПУ', 'Пангодинское ЛПУ'
                      'Пелымское ЛПУ', 'Перегребненское ЛПУ', 'Правохеттинское ЛПУ', 'Приозерное ЛПУ', 'Пунгинское ЛПУ', 'Сорумское ЛПУ', 'Сосновское ЛПУ',
                      'Таежное ЛПУ', 'Уральское ЛПУ', 'Ягельное ЛПУ', 'Ямбургское ЛПУ', 'Санаторий-профилакторий', 'КСК Норд'
                      ]

            # Создаем объект DataValidation
            data_validation = DataValidation(type="list", formula1='"{}"'.format(','.join(values)))

            # Применяем DataValidation к нужным ячейкам (например, H3 и I3)
            worksheet1.add_data_validation(data_validation)
            data_validation.add(worksheet1['D3'])

            values2 = ['За Декабрь 2023г', 'За Январь 2023г', 'За Февраль 2023г', 'За Март 2023г', 'За Апрель 2023г', 'За Май 2023г',
                      'За Июнь 2023г', 'За Июль 2023г', 'За Август 2023г', 'За Сентябрь 2023г',
                      'За Октябрь 2023г', 'За Ноябрь 2023г',
                      ]

            # Создаем объект DataValidation
            data_validation = DataValidation(type="list", formula1='"{}"'.format(','.join(values2)))

            # Применяем DataValidation к нужным ячейкам (например, H3 и I3)
            worksheet1.add_data_validation(data_validation)
            data_validation.add(worksheet1['D7'])

            # Запись значения по умолчанию в ячейку B2


            # Сохранение файла Excel
            table_end_row = worksheet1.max_row



            worksheet1.delete_cols(5, 3)  # Удаление столбиков с индексами 5, 6 и 7
            worksheet1.insert_rows(1, 12)  # Опускание таблицы на 12 строк ниже начиная с первой строки


            last_row = worksheet1.max_row
            sum_formula = f"=SUM(G1:G{last_row})"
            worksheet1.cell(row=last_row + 1, column=7).value = sum_formula

            last_row1 = worksheet1.max_row
            sum_formula1 = f"=SUM(H1:H{last_row})"
            worksheet1.cell(row=last_row1 + 0, column=8).value = sum_formula1

            last_row2 = worksheet1.max_row
            sum_formula1 = " "
            worksheet1.cell(row=last_row2 + 0, column=9).value = sum_formula1

            last_row2 = worksheet1.max_row
            sum_formula1 = "Итого:"
            worksheet1.cell(row=last_row2 + 0, column=6).value = sum_formula1


            worksheet1.cell(row=1, column=4).value = "Сводный реестр контрольных процедур"
            worksheet1.cell(row=2, column=4).value = " "
            worksheet1.cell(row=3, column=4).value = "_____________________________________"
            worksheet1.cell(row=4, column=4).value = "наименование филиала/отдела)"
            worksheet1.cell(row=6, column=4).value = "осуществляемых в целях налогового мониторинга"
            worksheet1.cell(row=7, column=4).value = "________________________________"


            #worksheet1.cell(row=table_end_row + 15, column=2).value = "________________________________"
            worksheet1.cell(row=table_end_row + 15, column=4).value = "________________________________"
            worksheet1.cell(row=table_end_row + 15, column=6).value = "________________________________"

            #worksheet1.cell(row=table_end_row + 16, column=2).value = "должность"
            worksheet1.cell(row=table_end_row + 16, column=4).value = "подпись"
            cell = worksheet1.cell(row=table_end_row + 16, column=6)
            cell.value = "ФИО"
            cell.alignment = Alignment(horizontal='left')


            workbook.save("example.xlsx")


            # Автоматическое расширение столбцов для первог о листа
            for row in range(worksheet1.max_row, 0, -1):
                max_length = 0
                for cell in worksheet1[row]:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                worksheet1.row_dimensions[row].height = max_length

                for column_cells in worksheet1.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet1.column_dimensions[column_cells[0].column_letter].width = length

                # Установка выравнивания для каждой ячейки в строке
                for cell in worksheet1[row]:
                    cell.alignment = Alignment(horizontal='centerContinuous', vertical='center', wrap_text=True)

                # Установка выравнивания для каждой ячейки в строке



                # Установка выравнивания для каждой ячейки в строке
            # Получение объекта worksheet для второго листа
            worksheet2 = writer.sheets['Sheet2']
            # Опускание таблицы на 2 строк ниже начиная с первой строки
            worksheet2.insert_rows(1, 2)
            worksheet2.cell(row=2, column=5).value = "Реестр объектов контроля"



            # Определение последней строки для столбца I
            last_row = worksheet2.max_row
            sum_formula = f"=SUM(I1:I{last_row})"
            worksheet2.cell(row=last_row + 1, column=9).value = sum_formula

            last_row1 = worksheet2.max_row
            sum_formula1 = f"=SUM(H1:H{last_row})"
            worksheet2.cell(row=last_row1 + 0, column=8).value = sum_formula1

            last_row2 = worksheet2.max_row
            sum_formula1 = "Итого:"
            cell = worksheet2.cell(row=last_row2 + 0, column=7)
            cell.value = sum_formula1
            cell.alignment = Alignment(horizontal='right')

            # Автоматическое расширение столбцов для второго листа
            for column_cells in worksheet2.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet2.column_dimensions[column_cells[0].column_letter].width = length



                for cell in worksheet2[row]:
                    cell.alignment = Alignment(horizontal='centerContinuous', vertical='center', wrap_text=True)


            # Получение объекта worksheet для третьего листа
            worksheet3 = writer.sheets['Sheet3']

            values3 = ['За Декабрь 2023г', 'За Январь 2023г', 'За Февраль 2023г', 'За Март 2023г', 'За Апрель 2023г', 'За Май 2023г',
                      'За Июнь 2023г', 'За Июль 2023г', 'За Август 2023г', 'За Сентябрь 2023г',
                      'За Октябрь 2023г', 'За Ноябрь 2023г',
                      ]

            # Создаем объект DataValidation
            data_validation = DataValidation(type="list", formula1='"{}"'.format(','.join(values3)))

            # Применяем DataValidation к нужным ячейкам (например, H3 и I3)
            worksheet3.add_data_validation(data_validation)
            data_validation.add(worksheet3['D5'])

            worksheet3.insert_rows(1, 7)

            # Удаление столбиков с индексами 5, 6 и 7


            worksheet3.delete_cols(3)


            worksheet3.delete_cols(4)
            worksheet3.delete_cols(4)


            worksheet3.delete_cols(5)
            worksheet3.delete_cols(5)

            worksheet3.delete_cols(4)

            worksheet3.cell(row=3, column=4).value = "осуществляемых в целях налогового мониторинга"

            last_row = worksheet3.max_row
            sum_formula = f"=SUM(D1:D{last_row})"
            worksheet3.cell(row=last_row + 1, column=4).value = sum_formula

            last_row1 = worksheet3.max_row
            sum_formula1 = f"=SUM(E1:E{last_row})"
            worksheet3.cell(row=last_row1 + 0, column=5).value = sum_formula1

            last_row2 = worksheet3.max_row
            sum_formula1 = "Итого:"
            cell = worksheet3.cell(row=last_row2 + 0, column=3)
            cell.value = sum_formula1
            cell.alignment = Alignment(horizontal='right')

            # После удаления столбка 4, столбок 6 станет столбком 7



            # Автоматическое расширение столбцов для второго листа
            for column_cells in worksheet3.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet3.column_dimensions[column_cells[0].column_letter].width = length

                for cell in worksheet3[row]:
                    cell.alignment = Alignment(horizontal='centerContinuous', vertical='center', wrap_text=True)





        file_path = 'summary.xlsx'  # Путь к файлу
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=summary.xlsx'
            return response

    return render(request, 'svod.html')  # Отображение шаблона "upload.html".


def base(request):
    if request.method == 'POST':
        # Получаем данные из формы, введенные в админке
        column1_data = request.POST.get('column1')
        column2_data = request.POST.get('column2')
        column3_data = request.POST.get('column3')
        column4_data = request.POST.get('column4')
        column5_data = request.POST.get('column5')
        column6_data = request.POST.get('column6')
        column7_data = request.POST.get('column7')
        column8_data = request.POST.get('column8')
        column9_data = request.POST.get('column9')
        column10_data = request.POST.get('column10')
        column11_data = request.POST.get('column11')
        column12_data = request.POST.get('column12')
        column13_data = request.POST.get('column13')
        column14_data = request.POST.get('column14')
        column15_data = request.POST.get('column15')

        # Создаем новый объект модели Columns с введенными данными
        new_column = Columns(column1=column1_data, column2=column2_data, column3=column3_data,
                             column4=column4_data, column5=column5_data, column6=column6_data,
                             column7=column7_data, column8=column8_data, column9=column9_data,
                             column10=column10_data,
                             column11=column11_data, column12=column12_data, column13=column13_data, column14=column14_data, column15=column15_data,
                             )
        new_column.save()

    # Получаем все объекты модели Columns
    columns = Columns.objects.all()
    return render(request, 'tables.html', {'columns': columns})


def checklist_detail(request, pk):
    checklist = get_object_or_404(CheckList, pk=pk)
    columns = CheckList.objects.all()
    if request.method == 'POST':
        form = CheckListForm(request.POST, instance=checklist)
        if form.is_valid():
            form.save()
            return redirect('checklist_detail', pk=checklist.pk)
    else:
        form = CheckListForm(instance=checklist)
    # Создание нового файла Excel
    workbook = Workbook()
    worksheet = workbook.active
    # Запись заголовков таблицы
    headers = ['номер п/п', 'Код КП(общий)', 'Код КП(промежуточный)', 'Наименование ИП', 'Описание КП', 'Переодичность проведения',
               'Способ подсчета результаты проведения КП', 'Подразделение, ответственное за проведение контрольной процедуры',
               'Исполнитель КП', 'Количество выполненых КП', 'Количество выявленных ошибок']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
    # Запись данных из базы данных в таблицу
    for row_num, column in enumerate(columns, 2):
        worksheet.cell(row=row_num, column=1).value = column.number
        worksheet.cell(row=row_num, column=2).value = column.cod_kp_overall
        worksheet.cell(row=row_num, column=3).value = column.cod_kp_intervall
        worksheet.cell(row=row_num, column=4).value = column.name_ip
        worksheet.cell(row=row_num, column=5).value = column.description_ip
        worksheet.cell(row=row_num, column=6).value = column.pereodiction_carriage
        worksheet.cell(row=row_num, column=7).value = column.counting_abillity
        worksheet.cell(row=row_num, column=8).value = column.responsible_group
        worksheet.cell(row=row_num, column=9).value = column.perforemr_kp
        worksheet.cell(row=row_num, column=10).value = column.number_complete
        worksheet.cell(row=row_num, column=11).value = column.number_mistakes
#        worksheet.cell(row=row_num, column=12).value = column.data_object
    # Сохранение файла Excel



    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=checklist.xlsx'
    workbook.save(response)
    return response


def edit_reestr(request, pk):
    reestr = get_object_or_404(Reestr, pk=pk)
    columns = Reestr.objects.all()  # Получение всех объектов Column
    if request.method == 'POST':
        form = ReestrForm(request.POST, instance=reestr)
        if form.is_valid():
            form.save()
            return redirect('edit_reestr', pk=reestr.pk)
    else:
        form = ReestrForm(instance=reestr)
    return render(request, 'reestr.html', {'reestr': reestr,'form': form, 'columns': columns })


def indexx(request):
    if request.method == 'POST':
        num_rows = int(request.POST['num_rows'])
        columns = []

        for i in range(num_rows):
            column = {
                'num': request.POST.get(f'num_{i}'),
                'cod_kp_inter': request.POST.get(f'cod_kp_inter_{i}'),
                'chek_num': request.POST.get(f'chek_num_{i}'),
                'obj_control': request.POST.get(f'obj_control_{i}'),
                'date_document': request.POST.get(f'date_document_{i}'),
                'num_document': request.POST.get(f'num_document_{i}'),
                'colvo_doc': request.POST.get(f'colvo_doc_{i}'),
                'colvo_errors': request.POST.get(f'colvo_errors_{i}'),
                'notes': request.POST.get(f'notes_{i}')
            }
            columns.append(column)

        # Создаем новый Excel-файл и заполняем его данными из таблицы
        workbook = Workbook()
        sheet = workbook.active



def download_excel(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from openpyxl.styles import Alignment
    from openpyxl import Workbook

    checklist1 = get_object_or_404(CheckList, pk=pk)
    checklist2 = get_object_or_404(Reestr, pk=pk)

    workbook = Workbook()
    # Запись данных из базы данных в таблицу
    worksheet1 = workbook.active
    # Задаем список значений для выпадающего списка

    # Задаем список значений для выпадающего списка
    values = [ 'АУП', 'Югорское УМТС и К', 'УОВОФ', 'Надымское УАВР', 'Югорское УАВР', 'Белоярское УАВР',
              'Надымское УТТиСТ', 'Югорское УТТиСТ', 'Белоярское УТТиСТ', 'ИТЦ',
              'Учебно-производственный центр', 'УЭЗ и С', 'Управление связи', 'Бобровское ЛПУ',
              'Верхнеказымское ЛПУ', 'Ивдельское ЛПУ', 'Казымское ЛПУ',
              'Карпинское ЛПУ', 'Комсомольское ЛПУ', 'Краснотурьинское ЛПУ',


              ]

    # Создаем объект DataValidation
    data_validation = DataValidation(type="list", formula1='"{}"'.format(','.join(values)))

    # Применяем DataValidation к нужным ячейкам (например, H3 и I3)
    worksheet1.add_data_validation(data_validation)
    data_validation.add(worksheet1['H3'])
    #data_validation.add(worksheet1['I3'])

    workbook.save('example.xlsx')

    worksheet1.cell(row=9, column=1).value = checklist1.number
    worksheet1.merge_cells('A8')
    worksheet1['A8'] = 'номер п/п'
    worksheet1['A8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    worksheet1.cell(row=9, column=2).value = checklist1.cod_kp_overall
    worksheet1.merge_cells('B8')
    worksheet1['B8'] = 'Код КП(общий)'
    worksheet1['B8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)


    worksheet1.cell(row=9, column=3).value = checklist1.cod_kp_intervall
    worksheet1.merge_cells('C8')
    worksheet1['C8'] = 'Код КП(Промежуточный)'
    worksheet1['C8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)


    worksheet1.cell(row=9, column=4).value = checklist1.name_ip
    worksheet1.merge_cells('D5')
    worksheet1['D8'] = 'Наименования КП'
    worksheet1['D8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)


    worksheet1.cell(row=9, column=5).value = checklist1.description_ip
    worksheet1.merge_cells('E8')
    worksheet1['E8'] = 'Описание КП'
    worksheet1['E8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    worksheet1.cell(row=9, column=6).value = checklist1.pereodiction_carriage
    worksheet1.merge_cells('F8')
    worksheet1['F8'] = 'Периодичность проведения (ежедневно/ ежеквартально/ежемесячно/по мере поступления и т.д)'
    worksheet1['F8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    worksheet1.cell(row=9, column=7).value = checklist1.counting_abillity
    worksheet1.merge_cells('G8')
    worksheet1['G8'] = 'Способ подсчета результаты  проведения КП (ручной/автоматизированный)'
    worksheet1['G8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    worksheet1.cell(row=9, column=8).value = checklist1.responsible_group
    worksheet1.merge_cells('H8')
    worksheet1['H8'] = 'Подразделение, ответственное за выполнение контрольной процедуры'
    worksheet1['H8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    worksheet1.cell(row=9, column=9).value = checklist1.perforemr_kp
    worksheet1.merge_cells('I8')
    worksheet1['I8'] = 'Исполнитель КП (ФИО)'
    worksheet1['I8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

#    worksheet1.cell(row=9, column=10).value = checklist1.number_complete   =Sheet2!H24
    worksheet1.cell(row=9, column=10).value = "=Sheet2!H24"
    worksheet1.merge_cells('J8')
    worksheet1['J8'] = 'Количество выполненных КП'
    worksheet1['J8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

#    worksheet1.cell(row=9, column=11).value = checklist1.number_mistakes  =Sheet2!I24
    worksheet1.cell(row=9, column=11).value = "=Sheet2!I24"
    worksheet1.merge_cells('K8')
    worksheet1['K8'] = 'Количество выявленных ошибок/ нарушений'
    worksheet1['K8'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    worksheet1.cell(row=3, column=7).value = checklist1.filial
    worksheet1.merge_cells('H3')
    worksheet1['H3'] = 'Филиал'
    worksheet1['H3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Создание стиля границы
    border_style = Border(left=Side(border_style="thin", color="000000"),
                          right=Side(border_style="thin", color="000000"),
                          top=Side(border_style="thin", color="000000"),
                          bottom=Side(border_style="thin", color="000000")
                          )

    # Автоматическое расширение столбцов
    for column in worksheet1.columns:
        max_length = 15
        column_letter = get_column_letter(column[8].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)

            except:
                pass
        adjusted_width = 20
        worksheet1.column_dimensions[column_letter].width = adjusted_width


    for row in worksheet1.rows:
        max_length = 1500
        for cell in row:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_height = 90
        alignment = Alignment(horizontal='centerContinuous', vertical='center', wrap_text=True)

        for cell in worksheet1[row[8].column]:
            worksheet1.row_dimensions[cell.row].height = adjusted_height
            cell.alignment = alignment

    # Сохранение файла
    workbook.save('example.xlsx')


    # Применение стиля границы к ячейкам
    for i in range(1, 8):
        for column in worksheet1.iter_cols(min_row=8, max_row=9, min_col=i, max_col=i + 4):
            for cell in column:
                cell.border = border_style


    worksheet1.cell(row=14, column=2).value = "_____________________"
    worksheet1.cell(row=15, column=2).value = "            (должность)"
    worksheet1.cell(row=14, column=4).value = "_____________________"
    worksheet1.cell(row=15, column=4).value = "             (подпись)"
    worksheet1.cell(row=14, column=6).value = "_____________________"
    worksheet1.cell(row=15, column=6).value = "                (ФИО)"
    worksheet1.cell(row=14, column=8).value = "______________________"
    worksheet1.cell(row=15, column=8).value = "                (дата)"
    #worksheet1.cell(row=3, column=8).value = "_____________________________________"
    worksheet1.cell(row=4, column=8).value = " (наименование филиала)"
    worksheet1.cell(row=2, column=11).value = "_____________________"
    worksheet1.cell(row=3, column=11).value = "   Код отдела/службы"
    worksheet1.cell(row=7, column=5).value = "          Чек-лист за"
    worksheet1.cell(row=7, column=5).font = worksheet1.cell(row=7, column=5).font.copy(bold=True)
    worksheet1.cell(row=7, column=6).value = " "
    worksheet1.cell(row=7, column=6).font = worksheet1.cell(row=7, column=6).font.copy(bold=True)
    worksheet1.cell(row=7, column=7).value = "            2023г."
    worksheet1.cell(row=7, column=7).font = worksheet1.cell(row=7, column=7).font.copy(bold=True)
    worksheet1.cell(row=3, column=8).value = f"=L9"


    worksheet2 = workbook.create_sheet(title='Sheet2')

    # Запись заголовков таблицы
    headers2 = ['№ п/п', 'Код КП(промежуточный)', 'Исполнитель ИП', 'номер чек листа',
               'Объект контроля (договор, акт, счет-фактура, КС-2 и др.)', 'Дата документа',
               'Номер документа',
               'Количество документов/операций',
               'Количество ошибок/нарушений',
               'Примечание',

               ]
    for col_num, header in enumerate(headers2, 1):
        cell = worksheet2.cell(row=3, column=col_num)
        cell.value = header

    # Запись данных из базы данных в таблицу
    worksheet2.cell(row=4, column=1).value = checklist2.num

    worksheet2.cell(row=4, column=2).value = checklist2.cod_kp_inter
    #    worksheet.merge_cells('B3')
    #    worksheet['B3'] = 'Код КП(промежуточный)'
    #    worksheet['B3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.alignment = Alignment(wrap_text=True)
    worksheet2.cell(row=4, column=3).value ="=Sheet!I9"

    worksheet2.cell(row=4, column=4).value = checklist2.chek_num
    worksheet2.cell(row=4, column=5).value = checklist2.obj_control
    worksheet2.cell(row=4, column=6).value = checklist2.date_document
    worksheet2.cell(row=4, column=7).value = checklist2.num_document
    worksheet2.cell(row=4, column=8).value = checklist2.colvo_doc


    last_row1 = worksheet2.max_row
    sum_formula1 = f"=SUM(H1:H23)"
    worksheet2.cell(row=last_row1 + 20, column=8).value = sum_formula1
    worksheet2.cell(row=4, column=9).value = checklist2.colvo_errors

    # Задаем формулу суммирования
    last_row = worksheet2.max_row
    sum_formula = f"=SUM(I1:I23)"
    worksheet2.cell(row=last_row , column=9).value = sum_formula
    worksheet2.cell(row=4, column=8).value = checklist2.notes

    total_errors = 0  # Инициализация переменной для суммирования ошибок

    border_style = Border(left=Side(border_style="thin", color="000000"),
                          right=Side(border_style="thin", color="000000"),
                          top=Side(border_style="thin", color="000000"),
#                          bottom=Side(border_style="thin", color="000000")
                          )

    # Автоматическое расширение столбцов
    for column in worksheet2.columns:
        max_length = 0
        column_letter = get_column_letter(column[3].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet2.column_dimensions[column_letter].width = adjusted_width

    # Автоматическое расширение строк

    # Применение стиля границы к ячейкам
    for i in range(1, 8):
        for column in worksheet2.iter_cols(min_row=3, max_row=25, min_col=i, max_col=i + 3):
            for cell in column:
                cell.border = border_style

    # Установка значений для "Должность" и "Подпись"
    worksheet2.cell(row=2, column=5).value = "                      Реестр обьектов контроля"
    worksheet2.cell(row=2, column=5).font = worksheet2.cell(row=2, column=5).font.copy(bold=True)
    worksheet2.cell(row=28, column=2).value = "_________________________"
    worksheet2.cell(row=29, column=2).value = "               (должность)"
    worksheet2.cell(row=28, column=4).value = "_______________________"
    worksheet2.cell(row=29, column=4).value = "               (ФИО)"
    worksheet2.cell(row=28, column=6).value = "_________________"
    worksheet2.cell(row=29, column=6).value = "         (подпись)"


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=checklist.xlsx'
    workbook.save(response)

    return response



def on_change(worksheet1, worksheet2):
    worksheet2.cell(row=4, column=3).value = worksheet1.cell(row=4, column=3).value

# Вызов функции on_change при изменении значения в ячейке worksheet1.cell(row=4, column=3)
    worksheet1.cell(row=4, column=3).add_observer(on_change, worksheet1, worksheet2)


def download_excel1(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    checklist = get_object_or_404(Reestr, pk=pk)

    workbook = Workbook()
    worksheet = workbook.active

    # Запись заголовков таблицы
    headers = ['№ п/п', 'Код КП(промежуточный)', 'Исполнитель ИП', 'номер чек листа', 'Объект контроля (договор, акт, счет-фактура, КС-2 и др.)', 'Дата документа',
               'Номер документа',
               'Количество документов/операций',
               'Количество ошибок/нарушений',
               'Примечание',
               ]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = header

    # Запись данных из базы данных в таблицу
    worksheet.cell(row=4, column=1).value = checklist.num


    worksheet.cell(row=4, column=2).value = checklist.cod_kp_inter
#    worksheet.merge_cells('B3')
#    worksheet['B3'] = 'Код КП(промежуточный)'
#    worksheet['B3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.alignment = Alignment(wrap_text=True)
    worksheet.cell(row=4, column=3).value = checklist.performer_ip
    worksheet.cell(row=4, column=4).value = checklist.chek_num
    worksheet.cell(row=4, column=5).value = checklist.obj_control
    worksheet.cell(row=4, column=6).value = checklist.date_document
    worksheet.cell(row=4, column=7).value = checklist.num_document
    worksheet.cell(row=4, column=8).value = checklist.colvo_doc
    worksheet.cell(row=4, column=9).value = checklist.colvo_errors
    worksheet.cell(row=4, column=10).value = checklist.notes

    total_errors = 0  # Инициализация переменной для суммирования ошибок

    border_style = Border(left=Side(border_style="thin", color="000000"),
                          right=Side(border_style="thin", color="000000"),
                          top=Side(border_style="thin", color="000000"),
                          bottom=Side(border_style="thin", color="000000")
                          )

    # Автоматическое расширение столбцов
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[3].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Автоматическое расширение строк


    # Применение стиля границы к ячейкам
    for i in range(1, 8):
        for column in worksheet.iter_cols(min_row=3, max_row=25, min_col=i, max_col=i + 3):
            for cell in column:
                cell.border = border_style




    # Установка значений для "Должность" и "Подпись"
    worksheet.cell(row=2, column=5).value = "                      Реестр обьектов контроля"
    worksheet.cell(row=2, column=5).font = worksheet.cell(row=2, column=5).font.copy(bold=True)
    worksheet.cell(row=28, column=2).value = "_________________________"
    worksheet.cell(row=29, column=2).value = "               (должность)"
    worksheet.cell(row=28, column=4).value = "_______________________"
    worksheet.cell(row=29, column=4).value = "               (ФИО)"
    worksheet.cell(row=28, column=6).value = "_________________"
    worksheet.cell(row=29, column=6).value = "         (подпись)"


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=checklist.xlsx'
    workbook.save(response)
    return response


def svod2(request):
    if request.method == 'POST':
        files = request.FILES.getlist('files')  # Получение списка загруженных файлов


        # Создание пустого сводного DataFrame для первого листа
        summary_df1 = pd.DataFrame(
            columns=['номер п/п', 'Код КП(общий)', 'Код КП(промежуточный)', 'Наименование ИП',
                     'Подразделение, ответственное за проведение контрольной процедуры', 'Исполнитель КП',
                     'Количество выполненых КП', 'Количество выявленных ошибок','Документ', 'Филиал'])



        # Обработка каждого загруженного файла
        for file in files:
            # Чтение первого листа файла и взятие только значений
            df1 = pd.read_excel(file, sheet_name='Sheet', usecols="A:J", header=None, skiprows=13, nrows=100)
            df1 = df1.set_axis(['номер п/п', 'Код КП(общий)', 'Код КП(промежуточный)', 'Наименование ИП',
                     'Подразделение, ответственное за проведение контрольной процедуры', 'Исполнитель КП',
                     'Количество выполненых КП', 'Количество выявленных ошибок','Документ', 'Филиал'], axis=1)

            # Получение имени файла без расширения
            file_name = os.path.splitext(file.name)[0]

            # Добавление столбика "Документ" в DataFrame и заполнение его названием файла
            df1['Документ'] = file_name

            df1.drop(['Документ'], axis=1,
                     inplace=True)
            df1 = df1.reset_index(drop=True)
            df1 = df1.rename_axis([None], axis=1)
            summary_df1 = pd.concat([summary_df1, df1], ignore_index=True)

            # Чтение второго листа файла и взятие только значений







        # Создание нового файла Excel с двумя листами
        with pd.ExcelWriter('summary.xlsx', engine='openpyxl') as writer:
            summary_df1.to_excel(writer, sheet_name='Sheet', index=False)


            # Получение объекта workbook
            workbook = writer.book

            # Получение объекта worksheet для первого листа
            worksheet1 = writer.sheets['Sheet']




            # Установка границ объединения ячеек

            worksheet1 = workbook.active

            #worksheet1.merge_cells('E14:E30')
            num_rows = summary_df1.shape[0]




            values = ['АУП', 'Югорское УМТС и К', 'УОВОФ', 'Надымское УАВР', 'Югорское УАВР', 'Белоярское УАВР',
                      'Надымское УТТиСТ', 'Югорское УТТиСТ', 'Белоярское УТТиСТ', 'ИТЦ',
                      'Учебно-производственный центр', 'УЭЗ и С', 'Управление связи', 'Бобровское ЛПУ',
                      'Верхнеказымское ЛПУ', 'Ивдельское ЛПУ', 'Казымское ЛПУ',
                      'Карпинское ЛПУ', 'Комсомольское ЛПУ', 'Краснотурьинское ЛПУ',
                      'Лонг-Юганское ЛПУ', 'Надымское ЛПУ', 'Нижнетуринское ЛПУ', 'Ново-Уренгойское ЛПУ',
                      'Ныдинское ЛПУ', 'Октябрьское ЛПУ', 'Пангодинское ЛПУ'
                      'Пелымское ЛПУ', 'Перегребненское ЛПУ', 'Правохеттинское ЛПУ', 'Приозерное ЛПУ',
                      'Пунгинское ЛПУ', 'Сорумское ЛПУ', 'Сосновское ЛПУ',
                      'Таежное ЛПУ', 'Уральское ЛПУ', 'Ягельное ЛПУ', 'Ямбургское ЛПУ', 'Санаторий-профилакторий', 'КСК Норд'
                      ]

            # Создаем объект DataValidation
            data_validation = DataValidation(type="list", formula1='"{}"'.format(','.join(values)))

            # Применяем DataValidation к нужным ячейкам (например, H3 и I3)
            worksheet1.add_data_validation(data_validation)
            data_validation.add(worksheet1['D3'])

            values2 = ['За Декабрь 2023г', 'За Январь 2023г', 'За Февраль 2023г', 'За Март 2023г', 'За Апрель 2023г', 'За Май 2023г',
                      'За Июнь 2023г', 'За Июль 2023г', 'За Август 2023г', 'За Сентябрь 2023г',
                      'За Октябрь 2023г', 'За Ноябрь 2023г',
                      ]

            # Создаем объект DataValidation
            data_validation = DataValidation(type="list", formula1='"{}"'.format(','.join(values2)))

            # Применяем DataValidation к нужным ячейкам (например, H3 и I3)
            worksheet1.add_data_validation(data_validation)
            data_validation.add(worksheet1['D7'])

            # Запись значения по умолчанию в ячейку B2


            # Сохранение файла Excel
            table_end_row = worksheet1.max_row


            #worksheet1.delete_cols(5, 3)
            worksheet1.delete_cols(3)
            worksheet1.delete_cols(3)
            worksheet1.delete_cols(3)
            worksheet1.delete_cols(3)
            worksheet1.delete_cols(5)




            #worksheet1.delete_cols(4)

            # Удаление столбиков с индексами 5, 6 и 7
            worksheet1.insert_rows(1, 12)

            for row in range(worksheet1.max_row, 0, -1):
                max_length = 20
                for cell in worksheet1[row]:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                worksheet1.row_dimensions[row].height = max_length

                for column_cells in worksheet1.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet1.column_dimensions[column_cells[0].column_letter].width = length

                # Установка выравнивания для каждой ячейки в строке
                for cell in worksheet1[row]:
                    cell.alignment = Alignment( vertical='center', wrap_text=True)
            # Опускание таблицы на 12 строк ниже начиная с первой строки







            workbook.save("example2.xlsx")


            # Автоматическое расширение столбцов для первого листа






        file_path = 'summary.xlsx'  # Путь к файлу
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=summary.xlsx'
            return response

    return render(request, 'svod2.html')  # Отображение шаблона "upload.html".